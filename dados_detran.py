import csv
from collections import Counter
from unidecode import unidecode
import os
from openpyxl import Workbook
arquivo_excel = Workbook()
tipos_veiculos = list()
tipos_infracoes = list()
tipo_arquivo = list()
def relatorios(pasta):
    arquivos = os.listdir(pasta)
    arquivos.sort()
    print(arquivos)
    return arquivos


def criar_planilha(contador_de_veiculos, contador_de_infracoes, arquivo):
    contador = 1
    planilha1 = arquivo_excel.create_sheet(arquivo)
    planilha1.cell(row=1, column=1, value='tipos de veiculos')
    planilha1.cell(row=1, column=2, value='Multas')
    for tipo_v, valor in contador_de_veiculos.items():
        print(f"{tipo_v} - {valor}")
        contador += 1
        planilha1.cell(row=contador, column=1, value=tipo_v)
        planilha1.cell(row=contador, column=2, value=valor)
    print('-=' * 50)
    for coluna, v in contador_de_infracoes.items():
        print(f"{coluna} - {v}")
    print('''
        
    ''')


def obter_tipo_infracao(linha):
    tipo_infracao = linha[12]
    return tipo_infracao


def obter_tipo_de_veiculo(linha):
    tipo_veiculo = unidecode(linha[3].lower().strip())
    return tipo_veiculo


for arquivo in relatorios('/Users/aman.rathie/PycharmProjects/detrans/relatorios'):
    print()
    print('==' * 50)
    print(arquivo)
    print('==' * 50)
    nome_arquivo = 'relatorios/' + arquivo
    with open(nome_arquivo) as csvfile:
        csv_file = csv.reader(csvfile, delimiter=';')
        for linha in csv_file:
            tipo_infracao = obter_tipo_infracao(linha)
            tipo_veiculo = obter_tipo_de_veiculo(linha)
            tipos_veiculos.append(tipo_veiculo)
            tipos_infracoes.append(tipo_infracao)
    contador_de_veiculos = Counter(tipos_veiculos)
    contador_de_infracoes = Counter(tipos_infracoes)
    del contador_de_veiculos['tipo_veiculo']
    criar_planilha(contador_de_veiculos, contador_de_infracoes, arquivo)

del arquivo_excel['Sheet']
arquivo_excel.save("relatorio.xlsx")