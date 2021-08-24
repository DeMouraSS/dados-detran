from openpyxl import Workbook
arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = "Relatorios"
planilha2 = arquivo_excel.create_sheet("Ganhos")
planilha1['A1'] = 'Categoria'
planilha1['B1'] = 'Valor'
planilha1['A2'] = "Restaurante"
planilha1['B2'] = 45.99
planilha2.cell(row=3, column=1, value=2000)
planilha2.cell(row=1, column=1, value="VALOR")
arquivo_excel.save("relatorio.xlsx")